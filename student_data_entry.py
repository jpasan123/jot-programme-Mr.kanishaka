import pandas as pd
import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tkcalendar import DateEntry  # Ensure you have tkcalendar installed
from PIL import Image, ImageTk

class StudentDataApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Job Oriented Training Programme Students & Data Entry")

        # Set window background color
        self.root.configure(bg='#f0f0f0')

        # Create a canvas and a scrollbar
        self.canvas = tk.Canvas(root, bg='#f0f0f0')
        self.scrollbar = tk.Scrollbar(root, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = tk.Frame(self.canvas, bg='#f0f0f0')

        # Pack the widgets
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        # Load and place logo
        self.logo_image = Image.open("kanishka_logo.png")
        self.logo_image = self.logo_image.resize((200, 100), Image.Resampling.LANCZOS)
        self.logo_photo = ImageTk.PhotoImage(self.logo_image)
        self.logo_label = tk.Label(self.scrollable_frame, image=self.logo_photo, bg='#f0f0f0')
        self.logo_label.grid(row=0, column=0, columnspan=2, pady=10)

        # Label and Entry widget configuration
        label_font = ("Arial", 12, "bold")
        entry_font = ("Arial", 12)
        entry_width = 30
        label_bg = '#f0f0f0'
        label_fg = '#333333'

        # Helper function to create labels and entries
        def create_label_entry(row, text, column=0):
            label = tk.Label(self.scrollable_frame, text=text, font=label_font, bg=label_bg, fg=label_fg)
            label.grid(row=row, column=column, sticky='e', padx=10, pady=5)
            entry = tk.Entry(self.scrollable_frame, font=entry_font, width=entry_width)
            entry.grid(row=row, column=column + 1, pady=5)
            entry.bind("<Return>", self.focus_next_widget)
            return entry

        # Helper function to create dropdown menus
        def create_dropdown(row, text, options, column=0):
            label = tk.Label(self.scrollable_frame, text=text, font=label_font, bg=label_bg, fg=label_fg)
            label.grid(row=row, column=column, sticky='e', padx=10, pady=5)
            var = tk.StringVar()
            dropdown = tk.OptionMenu(self.scrollable_frame, var, *options)
            dropdown.config(font=entry_font, width=entry_width - 5)
            dropdown.grid(row=row, column=column + 1, pady=5)
            var.set(options[0])
            return var
        
        # Section Titles
        section_title_font = ("Arial", 14, "bold")
        section_title_fg = 'red'
        section_title_bg = 'red'

        # Student Information Section

         # Additional Fields Section
        
        tk.Label(self.scrollable_frame, text="Student & Modules & Lectures Information", font=section_title_font, bg=section_title_bg, fg=section_title_fg).grid(row=24, column=0, columnspan=2, pady=10)

        self.student_no_entry = create_label_entry(1, "Student No")
        self.student_name_entry = create_label_entry(2, "Student Name")
        self.program_name_entry = create_label_entry(3, "Program Name")
        self.jot_activity1_entry = create_label_entry(4, "JOT Activity")
        self.phicycle_activity1_entry = create_label_entry(5, "Phicycle Activity")
        self.lab_section1_entry = create_label_entry(6, "Lab Section")
        self.assignment1_entry = create_label_entry(7, "Assignment")
        self.assessment1_entry = create_label_entry(8, "Assessment")
        self.practical1_entry = create_label_entry(9, "Practical")
        self.module1_title_entry = create_label_entry(10, "Module Title")
        self.module1_subject1_entry = create_label_entry(11, "Module Subject")

        # Module 1 Hours and Minutes
        self.module1_hours_label = tk.Label(self.scrollable_frame, text="Module Hours", font=label_font, bg=label_bg, fg=label_fg)
        self.module1_hours_label.grid(row=12, column=0, sticky='e', padx=10, pady=5)
        self.module1_hours_spinbox = tk.Spinbox(self.scrollable_frame, from_=0, to=23, wrap=True, state="readonly", width=5, font=entry_font)
        self.module1_hours_spinbox.grid(row=12, column=1, sticky='w', pady=5)
        self.module1_hours_spinbox.bind("<Return>", self.focus_next_widget)
        self.module1_minutes_spinbox = tk.Spinbox(self.scrollable_frame, from_=0, to=59, wrap=True, state="readonly", width=5, font=entry_font)
        self.module1_minutes_spinbox.grid(row=12, column=1, sticky='e', pady=5)
        self.module1_minutes_spinbox.bind("<Return>", self.focus_next_widget)

        self.module1_supervisor_entry = create_label_entry(13, "Module Supervisor")
        self.module1_lecturer_entry = create_label_entry(14, "Module Lecturer")
        self.module1_satisfaction_entry = create_label_entry(15, "Lecturer or Supervisor Satisfaction")
        self.module1_lecturer_no_entry = create_label_entry(16, "Module Lecturer No")
        self.module1_supervisor_no_entry = create_label_entry(17, "Module Supervisor No")
        self.module1_title_no_entry = create_label_entry(18, "Module Title No")
        self.module1_subject_no_entry = create_label_entry(19, "Module Subject No")

        # Subject Hours and Minutes
        self.subject_hours_label = tk.Label(self.scrollable_frame, text="Subject Hours", font=label_font, bg=label_bg, fg=label_fg)
        self.subject_hours_label.grid(row=20, column=0, sticky='e', padx=10, pady=5)
        self.subject_hours_spinbox = tk.Spinbox(self.scrollable_frame, from_=0, to=23, wrap=True, state="readonly", width=5, font=entry_font)
        self.subject_hours_spinbox.grid(row=20, column=1, sticky='w', pady=5)
        self.subject_hours_spinbox.bind("<Return>", self.focus_next_widget)
        self.subject_minutes_spinbox = tk.Spinbox(self.scrollable_frame, from_=0, to=59, wrap=True, state="readonly", width=5, font=entry_font)
        self.subject_minutes_spinbox.grid(row=20, column=1, sticky='e', pady=5)
        self.subject_minutes_spinbox.bind("<Return>", self.focus_next_widget)

        self.module_date_label = tk.Label(self.scrollable_frame, text="Module Date", font=label_font, bg=label_bg, fg=label_fg)
        self.module_date_label.grid(row=21, column=0, sticky='e', padx=10, pady=5)
        self.module_date_entry = DateEntry(self.scrollable_frame, font=entry_font)
        self.module_date_entry.grid(row=21, column=1, pady=5)
        self.module_date_entry.bind("<Return>", self.focus_next_widget)

        self.subject_date_label = tk.Label(self.scrollable_frame, text="Subject Date", font=label_font, bg=label_bg, fg=label_fg)
        self.subject_date_label.grid(row=22, column=0, sticky='e', padx=10, pady=5)
        self.subject_date_entry = DateEntry(self.scrollable_frame, font=entry_font)
        self.subject_date_entry.grid(row=22, column=1, pady=5)
        self.subject_date_entry.bind("<Return>", self.focus_next_widget)

        self.final_approval_entry = create_label_entry(23, "Final Approval Note")

        # Additional Fields Section
        tk.Label(self.scrollable_frame, text="Personal Information", font=section_title_font, fg='red', bg='#f0f0f0').grid(row=24, column=0, columnspan=2, pady=10)        
        self.full_name_entry = create_label_entry(25, "Name In Full")
        self.initial_name_entry = create_label_entry(26, "Name With Initial")
        self.dob_label = tk.Label(self.scrollable_frame, text="Date Of Birth", font=label_font, bg=label_bg, fg=label_fg)
        self.dob_label.grid(row=28, column=0, sticky='e', padx=10, pady=5)
        self.dob_entry = DateEntry(self.scrollable_frame, width=entry_width - 3, font=entry_font)
        self.dob_entry.grid(row=28, column=1, pady=5)
        self.dob_entry.bind("<Return>", self.focus_next_widget)

        self.nic_entry = create_label_entry(28, "National Identity Card No")
        
        tk.Label(self.scrollable_frame, text="Gender", font=label_font, bg=label_bg, fg=label_fg).grid(row=29, column=0, sticky='e', padx=10, pady=5)
        self.gender_var = tk.StringVar()
        self.gender_dropdown = tk.OptionMenu(self.scrollable_frame, self.gender_var, "Male", "Female", "Rather Not Say")
        self.gender_dropdown.config(font=entry_font, width=entry_width - 5)
        self.gender_dropdown.grid(row=29, column=1, pady=5)
        self.gender_var.set("Select Gender")
        
        self.postal_address_entry = create_label_entry(30, "Postal Address")
        self.id_address_entry = create_label_entry(31, "Address In ID")
        self.permanent_address_entry = create_label_entry(32, "Permanent Address")
        self.current_address_entry = create_label_entry(33, "Current Address")
        self.mobile_number_entry = create_label_entry(34, "Mobile Number")
        self.land_phone_number_entry = create_label_entry(35, "Land Phone Number")
        self.whatsapp_number_entry = create_label_entry(36, "Whatsapp Number")
        self.facebook_url_entry = create_label_entry(37, "Facebook URL")
        self.instagram_id_entry = create_label_entry(38, "Instagram ID")
        self.email_entry = create_label_entry(39, "Email")

        # Geographical Details Section
        tk.Label(self.scrollable_frame, text="Geographical Details", font=section_title_font, fg=section_title_fg).grid(row=40, column=0, columnspan=2, pady=10)
        
        provinces = ["Select Province", "Central Province", "Eastern Province", "Northern Province", "Southern Province", "Western Province", "North Western Province", "North Central Province", "Uva Province", "Sabaragamuwa Province"]
        self.province_var = create_dropdown(41, "Province", provinces)
        districts = [
            "Select District", "Colombo", "Gampaha", "Kalutara", "Kandy", "Matale", "Nuwara Eliya", 
            "Galle", "Matara", "Hambantota", "Jaffna", "Kilinochchi", "Mannar", "Vavuniya", 
            "Mullaitivu", "Batticaloa", "Ampara", "Trincomalee", "Kurunegala", "Puttalam", 
            "Anuradhapura", "Polonnaruwa", "Badulla", "Moneragala", "Ratnapura", "Kegalle"
        ]
        self.district_var = create_dropdown(42, "District", districts)

        ds_divisions = [
            "Select DS Division", "Alawwa", "Ambanpola", "Bamunakotuwa", "Bingiriya", "Polgahawela", 
            "Polpithigama", "Pothuhera", "Ridigama", "Wariyapola", "Weerambugedara", "Kobeigane", 
            "Kuliyapitiya East", "Kuliyapitiya West", "Kurunegala", "Mahawa", "Mallawapitiya", 
            "Maspotha", "Mawathagama", "Narammala", "Panduwasnuwara East", "Panduwasnuwara West", 
            "Udubaddawa"
        ]
        self.ds_division_var = create_dropdown(43, "DS Division", ds_divisions)
        self.gn_division_entry = create_label_entry(44, "GN Division")
        self.nearest_city_entry = create_label_entry(45, "Nearest City")

        # Parent/Guardian Information Section
        tk.Label(self.scrollable_frame, text="Parent or Guardian Information", font=section_title_font, fg=section_title_fg).grid(row=46, column=0, columnspan=2, pady=10)
        
        self.guardian_name_entry = create_label_entry(47, "Full Name")
        self.guardian_dob_label = tk.Label(self.scrollable_frame, text="Date of Birth", font=label_font, bg=label_bg, fg=label_fg)
        self.guardian_dob_label.grid(row=48, column=0, sticky='e', padx=10, pady=5)
        self.guardian_dob_entry = DateEntry(self.scrollable_frame, width=entry_width - 3, font=entry_font)
        self.guardian_dob_entry.grid(row=48, column=1, pady=5)
        self.guardian_dob_entry.bind("<Return>", self.focus_next_widget)
        self.guardian_age_entry = create_label_entry(49, "Age")
        self.guardian_occupation_entry = create_label_entry(50, "Occupation")
        self.guardian_contact_number_entry = create_label_entry(51, "Contact Number")
        self.guardian_whatsapp_number_entry = create_label_entry(52, "WhatsApp Number")
        self.guardian_address_entry = create_label_entry(53, "Address")
        self.guardian_email_entry = create_label_entry(54, "Email")

        # Educational Qualifications Section
        tk.Label(self.scrollable_frame, text="Educational Qualifications", font=section_title_font, fg=section_title_fg).grid(row=55, column=0, columnspan=2, pady=10)
        
        # G.C.E (O/L)
        tk.Label(self.scrollable_frame, text="G.C.E (O/L) Examination", font=label_font, bg=label_bg, fg=label_fg).grid(row=56, column=0, columnspan=2, pady=5)
        self.ol_school_entry = create_label_entry(57, "Name of School/College")
        self.ol_admission_number_entry = create_label_entry(58, "Admission Number of Examination (O/L)")
        self.ol_year_label = tk.Label(self.scrollable_frame, text="Year of Sitting", font=label_font, bg=label_bg, fg=label_fg)
        self.ol_year_label.grid(row=59, column=0, sticky='e', padx=10, pady=5)
        self.ol_year_entry = DateEntry(self.scrollable_frame, width=entry_width - 3, font=entry_font)
        self.ol_year_entry.grid(row=59, column=1, pady=5)
        self.ol_year_entry.bind("<Return>", self.focus_next_widget)
        self.ol_award_achievements_entry = create_label_entry(60, "Award Achievements")

        # G.C.E (A/L)
        tk.Label(self.scrollable_frame, text="G.C.E (A/L) Examination", font=label_font, bg=label_bg, fg=label_fg).grid(row=61, column=0, columnspan=2, pady=5)
        self.al_school_entry = create_label_entry(62, "Name of School/College")
        self.al_admission_number_entry = create_label_entry(63, "Admission Number of Examination (A/L)")
        self.al_year_label = tk.Label(self.scrollable_frame, text="Year of Sitting", font=label_font, bg=label_bg, fg=label_fg)
        self.al_year_label.grid(row=59, column=0, sticky='e', padx=10, pady=5)
        self.al_year_entry = DateEntry(self.scrollable_frame, width=entry_width - 3, font=entry_font)
        self.al_year_entry.grid(row=59, column=1, pady=5)
        self.al_year_entry.bind("<Return>", self.focus_next_widget)
        self.al_award_achievements_entry = create_label_entry(65, "Award Achievements")

        # Professional Qualifications Section
        tk.Label(self.scrollable_frame, text="Professional Qualifications", font=section_title_font, fg=section_title_fg).grid(row=66, column=0, columnspan=2, pady=10)
        
        self.pro_institution_entry = create_label_entry(67, "Institution")
        self.pro_result_achievement_entry = create_label_entry(68, "Result Achievement")
        self.pro_pass_out_year_label = tk.Label(self.scrollable_frame, text="Pass Out Year", font=label_font, bg=label_bg, fg=label_fg)
        self.pro_pass_out_year_label.grid(row=69, column=0, sticky='e', padx=10, pady=5)
        self.pro_pass_out_year_entry = DateEntry(self.scrollable_frame, width=entry_width - 3, font=entry_font)
        self.pro_pass_out_year_entry.grid(row=69, column=1, pady=5)
        self.pro_pass_out_year_entry.bind("<Return>", self.focus_next_widget)
        self.pro_areas_covered_entry = create_label_entry(70, "Areas Covered")
        self.pro_any_other_entry = create_label_entry(71, "Any Other")

        # Other Available Qualifications Section
        tk.Label(self.scrollable_frame, text="Other Available Qualifications", font=section_title_font, fg=section_title_fg).grid(row=72, column=0, columnspan=2, pady=10)
        
        self.other_qualifications_entry = create_label_entry(73, "Other Qualifications")

        # Special Talents Section
        tk.Label(self.scrollable_frame, text="Special Talents", font=section_title_font, fg=section_title_fg).grid(row=74, column=0, columnspan=2, pady=10)
        
        self.special_talent_entry = create_label_entry(75, "Special Talent (e.g., Dancing, Singing, Writing, etc.)")

        # Employment Status Section
        tk.Label(self.scrollable_frame, text="Employment Status", font=section_title_font, fg=section_title_fg).grid(row=76, column=0, columnspan=2, pady=10)
        
        self.employment_institution_entry = create_label_entry(77, "Name of the Institution")
        self.employment_job_title_entry = create_label_entry(78, "Job Title")
        self.employment_joining_year_label = tk.Label(self.scrollable_frame, text="Joining Year", font=label_font, bg=label_bg, fg=label_fg)
        self.employment_joining_year_label.grid(row=79, column=0, sticky='e', padx=10, pady=5)
        self.employment_joining_year_entry = DateEntry(self.scrollable_frame, width=entry_width - 3, font=entry_font)
        self.employment_joining_year_entry.grid(row=79, column=1, pady=5)
        self.employment_joining_year_entry.bind("<Return>", self.focus_next_widget)
        self.employment_year_left_label = tk.Label(self.scrollable_frame, text="Year Left", font=label_font, bg=label_bg, fg=label_fg)
        self.employment_year_left_label.grid(row=80, column=0, sticky='e', padx=10, pady=5)
        self.employment_year_left_entry = DateEntry(self.scrollable_frame, width=entry_width - 3, font=entry_font)
        self.employment_year_left_entry.grid(row=80, column=1, pady=5)
        self.employment_year_left_entry.bind("<Return>", self.focus_next_widget)

        # Declaration Section
        tk.Label(self.scrollable_frame, text="Declaration", font=section_title_font, fg=section_title_fg).grid(row=81, column=0, columnspan=2, pady=10)
        
        declaration_text = "I, declare that the information provided herein is accurate, complete, and true to my knowledge. " \
                           "I understand that any false statement or omission may have legal implications and result in " \
                           "the denial or revocation of services, employment, or legal standings dependent on this declaration. " \
                           "By signing this declaration, I consent and agree to the above-mentioned terms concerning my biodata."
        
        self.declaration_label = tk.Label(self.scrollable_frame, text=declaration_text, font=("Arial", 10), bg=label_bg, fg=label_fg, wraplength=600, justify="left")
        self.declaration_label.grid(row=82, column=0, columnspan=2, padx=10, pady=10)
        
        self.agree_var = tk.IntVar()
        self.agree_checkbox = tk.Checkbutton(self.scrollable_frame, text="I Agree", variable=self.agree_var, font=entry_font, bg=label_bg)
        self.agree_checkbox.grid(row=83, column=0, columnspan=2, pady=5)

        # Submit Button
        self.submit_button = tk.Button(self.scrollable_frame, text="Submit", font=("Arial", 14, "bold"), bg="#4CAF50", fg="white", command=self.submit_data)
        self.submit_button.grid(row=84, column=0, columnspan=2, pady=10)

    def focus_next_widget(self, event):
        event.widget.tk_focusNext().focus()
        return "break"

    def submit_data(self):
        student_no = self.student_no_entry.get()
        student_name = self.student_name_entry.get()
        program_name = self.program_name_entry.get()
        jot_activity1 = self.jot_activity1_entry.get()
        phicycle_activity1 = self.phicycle_activity1_entry.get()
        lab_section1 = self.lab_section1_entry.get()
        assignment1 = self.assignment1_entry.get()
        assessment1 = self.assessment1_entry.get()
        practical1 = self.practical1_entry.get()
        module1_title = self.module1_title_entry.get()
        module1_subject1 = self.module1_subject1_entry.get()
        
        # Get values from Spinbox widgets
        module1_hours = f"{self.module1_hours_spinbox.get()}:{self.module1_minutes_spinbox.get()}"
        subject_hours = f"{self.subject_hours_spinbox.get()}:{self.subject_minutes_spinbox.get()}"
        
        module1_supervisor = self.module1_supervisor_entry.get()
        module1_lecturer = self.module1_lecturer_entry.get()
        module1_satisfaction = self.module1_satisfaction_entry.get()
        module1_lecturer_no = self.module1_lecturer_no_entry.get()
        module1_supervisor_no = self.module1_supervisor_no_entry.get()
        module1_title_no = self.module1_title_no_entry.get()
        module1_subject_no = self.module1_subject_no_entry.get()
        
        module_date = self.module_date_entry.get_date()
        subject_date = self.subject_date_entry.get_date()
        final_approval = self.final_approval_entry.get()

        # Additional Fields
        full_name = self.full_name_entry.get()
        initial_name = self.initial_name_entry.get()
        dob = self.dob_entry.get()
        nic = self.nic_entry.get()
        gender = self.gender_var.get()
        postal_address = self.postal_address_entry.get()
        id_address = self.id_address_entry.get()
        permanent_address = self.permanent_address_entry.get()
        current_address = self.current_address_entry.get()
        mobile_number = self.mobile_number_entry.get()
        land_phone_number = self.land_phone_number_entry.get()
        whatsapp_number = self.whatsapp_number_entry.get()
        facebook_url = self.facebook_url_entry.get()
        instagram_id = self.instagram_id_entry.get()
        email = self.email_entry.get()

        province = self.province_var.get()
        district = self.district_var.get()
        ds_division = self.ds_division_var.get()
        gn_division = self.gn_division_entry.get()
        nearest_city = self.nearest_city_entry.get()

        guardian_name = self.guardian_name_entry.get()
        guardian_dob = self.guardian_dob_entry.get()
        guardian_age = self.guardian_age_entry.get()
        guardian_occupation = self.guardian_occupation_entry.get()
        guardian_contact_number = self.guardian_contact_number_entry.get()
        guardian_whatsapp_number = self.guardian_whatsapp_number_entry.get()
        guardian_address = self.guardian_address_entry.get()
        guardian_email = self.guardian_email_entry.get()

        ol_school = self.ol_school_entry.get()
        ol_admission_number = self.ol_admission_number_entry.get()
        ol_year = self.ol_year_entry.get()
        ol_award_achievements = self.ol_award_achievements_entry.get()

        al_school = self.al_school_entry.get()
        al_admission_number = self.al_admission_number_entry.get()
        al_year = self.al_year_entry.get()
        al_award_achievements = self.al_award_achievements_entry.get()

        pro_institution = self.pro_institution_entry.get()
        pro_result_achievement = self.pro_result_achievement_entry.get()
        pro_pass_out_year = self.pro_pass_out_year_entry.get()
        pro_areas_covered = self.pro_areas_covered_entry.get()
        pro_any_other = self.pro_any_other_entry.get()

        other_qualifications = self.other_qualifications_entry.get()

        special_talent = self.special_talent_entry.get()

        employment_institution = self.employment_institution_entry.get()
        employment_job_title = self.employment_job_title_entry.get()
        employment_joining_year = self.employment_joining_year_entry.get()
        employment_year_left = self.employment_year_left_entry.get()

        agree = self.agree_var.get()

        if agree != 1:
            messagebox.showwarning("Warning", "You must agree to the declaration.")
            return

        # Save the data into an Excel file
        try:
            file_path = "student_data.xlsx"
            wb = load_workbook(file_path)
            ws = wb.active

            # Find the next empty row in the worksheet
            next_row = ws.max_row + 1

            # Write the data to the Excel sheet
            ws.append([
                student_no, student_name, program_name, jot_activity1, phicycle_activity1, lab_section1, assignment1, assessment1,
                practical1, module1_title, module1_subject1, module1_hours, subject_hours, module1_supervisor, module1_lecturer,
                module1_satisfaction, module1_lecturer_no, module1_supervisor_no, module1_title_no, module1_subject_no, module_date,
                subject_date, final_approval, full_name, initial_name, dob, nic, gender, postal_address, id_address, permanent_address,
                current_address, mobile_number, land_phone_number, whatsapp_number, facebook_url, instagram_id, email, province, 
                district, ds_division, gn_division, nearest_city, guardian_name, guardian_dob, guardian_age, guardian_occupation, 
                guardian_contact_number, guardian_whatsapp_number, guardian_address, guardian_email, ol_school, ol_admission_number, 
                ol_year, ol_award_achievements, al_school, al_admission_number, al_year, al_award_achievements, pro_institution, 
                pro_result_achievement, pro_pass_out_year, pro_areas_covered, pro_any_other, other_qualifications, special_talent, 
                employment_institution, employment_job_title, employment_joining_year, employment_year_left
            ])

            # Save the workbook
            wb.save(file_path)

            messagebox.showinfo("Success", "Data saved successfully.")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while saving the data: {str(e)}")

# Main application loop
if __name__ == "__main__":
    root = tk.Tk()
    app = StudentDataApp(root)
    root.mainloop()
